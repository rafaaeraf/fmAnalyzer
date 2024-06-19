import os
import random
import re
import sys

import numpy as np
import pandas as pd

from openpyxl import load_workbook

################## Inputs ##################
IN_GAME_DATE = "2030-06-10" # YYYY-MM-DD
DATA_DIR = "C:\\fmAnalyzer\\data"
OUTPUT_DIR = "C:\\fmAnalyzer\\output"
CLUB_NAME = "Blyth"
WEIGHTS_DICT_PLAYERS = {"green": 5, "blue": 3, "normal": 1}
WEIGHTS_DICT_COACHES = {"green": 5, "blue": 3, "normal": 0}
TEAM_FORMATION = ["GR", "D(D)", "D(C)", "D(C)", "D(E)", "MD", "M(C)", "M(C)", "MO(C)", "PL(C)",
                   "PL(C)"]

################## Methods ##################
def get_df_type(df):
    if "Posição" in df.keys():
        return "players"
    elif "Função" in df.keys():
        return "coaches"

def get_df_type_detailed(df):
    if get_df_type(df) == "players":
        # If all players are from the club or lent by the club, then this file is a
        # club export file
        # There is an exception when the player is out in experience. There is no way to
        # detect that. So we say that if almost all players are from the club, this must
        # be a club file
        if (sum(df["Clube"] == CLUB_NAME) + sum(df["Emprestado Por"] == CLUB_NAME)) / len(df) > 0.9:
            return "my_club_players"
        else:
            return "search_players"
    elif get_df_type(df) == "coaches":
        if all(df["Clube"] == CLUB_NAME):
            return "my_club_coaches"
        else:
            return "search_coaches"
    else:
        sys.exit("ERROR: Data is not for coaches neither players")

def read_data():
    all_files = []
    for data_file in os.listdir(DATA_DIR):
        file_full_path = os.path.join(DATA_DIR, data_file)
        all_files.append({"file": file_full_path, "timestamp": os.path.getmtime(file_full_path)})
    all_files_df = pd.DataFrame(all_files)
    all_files_df = all_files_df.sort_values("timestamp")

    # Create list of dataframes
    players_list = []
    coaches_list = []

    club_players_df = None
    club_coaches_df = None
    club_players_last_date = 0
    club_coaches_last_date = 0
    for file in all_files_df["file"]:
        # Read html
        tmp_df = pd.read_html(file, header=0, encoding="utf-8", na_values=["-", "- -"])
        if len(tmp_df) != 1:
            sys.exit("ERROR: All generated tables should have length 1")
        tmp_df = tmp_df[0]
        # Check if the source is missing names (FM bug)
        if tmp_df["Nome"].isnull().any():
            sys.exit('ERROR: Provided table "' + file + '" have empty names')
        # Add relevant info to df
        tmp_df["source_file"] = file
        tmp_df["timestamp"] = all_files_df[all_files_df["file"] == file]["timestamp"].values[0]
        # Get dataframe type
        df_type = get_df_type(tmp_df)
        tmp_df["df_type"] = df_type
        tmp_df["df_type_detailed"] = get_df_type_detailed(tmp_df)
        # We should have only one coaches and players df, to avoid having people that left
        # Keep the latest one and only append in the end
        if all(tmp_df["df_type_detailed"] == "my_club_players"):
            if os.path.getmtime(file) > club_players_last_date:
                club_players_last_date = os.path.getmtime(file)
                club_players_df = tmp_df
        elif all(tmp_df["df_type_detailed"] == "my_club_coaches"):
            if os.path.getmtime(file) > club_coaches_last_date:
                club_coaches_last_date = os.path.getmtime(file)
                club_coaches_df = tmp_df
        else:
            # Append search dfs to the list
            if df_type == "players":
                players_list.append(tmp_df)
            else:
                coaches_list.append(tmp_df)

    # After looking at all files, add the club ones to the list
    if club_players_df is not None:
        players_list.append(club_players_df)
    if club_coaches_df is not None:
        coaches_list.append(club_coaches_df)

    # Concatenate DataFrames outside the loop
    players_ret = pd.concat(players_list, axis=0, ignore_index=True)
    coaches_ret = pd.concat(coaches_list, axis=0, ignore_index=True)

    return players_ret, coaches_ret

def get_person_club_status(row):
    if row["df_type_detailed"] == "my_club_players":
        if row["Clube"] == CLUB_NAME:
            return "my_club"
        else:
            return "my_club_lent"
    elif row["df_type_detailed"] == "search_players":
        if pd.isna(row["Clube"]):
            return "has_no_club"
        else:
            return "has_club"
    elif row["df_type_detailed"] == "my_club_coaches":
        return "my_club"
    elif row["df_type_detailed"] == "search_coaches":
        if pd.isna(row["Clube"]):
            return "has_no_club"
        else:
            return "has_club"

def parse_position(orig_position):
    orig_position = orig_position.split(",")
    ret = []
    for o in orig_position:
        side = None
        if "(" in o:
            side = o.split("(")[1].split(")")[0]
        positions_internal = o.split("/")
        for p in positions_internal:
            curr_position = p.replace(" ", "").split("(")[0]
            if side is None:
                ret.append(curr_position)
            else:
                for s in side:
                    ret.append(curr_position + "(" + s + ")")
    return ",".join(ret)

def get_relevant_attributes(edited_positions, attributes_internal):
    if "GR" in edited_positions:
        return attributes_internal[(attributes_internal["gk_attribute"] == "both") |
                                   (attributes_internal["gk_attribute"] == "gk")]["short_name"]
    else:
        return attributes_internal[(attributes_internal["gk_attribute"] == "both") |
                                   (attributes_internal["gk_attribute"] == "non_gk")]["short_name"]

def check_data_completion(row, attributes_internal):
    if row["df_type"] == "coaches":
        relevant_attributes = attributes_internal.loc[
            attributes_internal["player_coach"] == "coach", "short_name"]
    else:
        edited_positions = row["edited_positions"]
        relevant_attributes = get_relevant_attributes(edited_positions, attributes_internal)
    filtered_df = row[relevant_attributes]
    if all(pd.isna(filtered_df.values)):
        return "all_incomplete"
    elif any(pd.isna(filtered_df.values)):
        return "incomplete"
    elif any(filtered_df.astype(str).str.contains("-")):
        return "completeButRange"
    else:
        return "complete"

def convert_df_columns_to_numeric(df, columns):
    df_copy = df.copy()
    for col in columns:
        df_copy[col] = pd.to_numeric(df_copy[col], errors="coerce")
    return df_copy

def get_all_possible_positions(players_internal):
    all_positions = [p.split(",") for p in players_internal["edited_positions"].unique()]
    all_positions = np.unique(np.concatenate(all_positions))
    # Get unique values from TEAM_FORMATION keeping the order and add new values found in players DB
    all_positions = list(dict.fromkeys(TEAM_FORMATION)) + [
        pos for pos in all_positions if pos not in TEAM_FORMATION]
    return all_positions

def filter_players_by_position(df, pos):
    return df[df["edited_positions"].str.contains(pos,regex=False)]

def get_average_attribute_per_position(players_internal, attributes_internal):
    all_positions = get_all_possible_positions(players_internal)
    ret = pd.DataFrame()
    complete_players = players_internal[players_internal["data_completion"] == "complete"]
    for pos in all_positions:
        relevant_attributes = get_relevant_attributes(pos, attributes_internal)
        filtered_df = filter_players_by_position(complete_players, pos)
        filtered_df = filtered_df[relevant_attributes]
        filtered_df = convert_df_columns_to_numeric(filtered_df, relevant_attributes)
        ret = pd.concat([ret, filtered_df.mean().to_frame().T], ignore_index=True)
    ret["positions"] = all_positions
    return ret

def handle_missing_attributes(row, attributes_internal, average_attributes_pos_internal):
    if row["data_completion"] in ["all_incomplete", "complete"]:
        return row
    row = row.copy()
    relevant_attributes = get_relevant_attributes(row['edited_positions'], attributes_internal)
    player_position = random.choice(row["edited_positions"].split(","))
    for attribute_name in relevant_attributes:
        if pd.isna(row[attribute_name]):
            selected_attribute = average_attributes_pos_internal.loc[
                average_attributes_pos_internal["positions"] == player_position, attribute_name]
            if len(selected_attribute) != 1:
                sys.exit("ERROR: Unexpected length for the selected_attribute")
            row.at[attribute_name] = float(selected_attribute.iloc[0])
        elif "-" in str(row[attribute_name]):
            range_attribute = row[attribute_name].split("-")
            row.at[attribute_name] = (float(range_attribute[0]) + float(range_attribute[1])) / 2
    return row

def filter_positions_df_by_position(positions_internal, pos):
    return positions_internal.loc[positions_internal["positions"].str.contains(re.escape(pos),
                                                                               na=False), ]

# Create new fake roles, one per position that represent that position in general
# Create an average of all weights for the roles on that positions to represent the general
# weights per position
def create_general_roles(players_internal, positions_internal, weights_internal):
    # Handle positions dataframe
    all_positions = get_all_possible_positions(players_internal)
    fake_roles = pd.DataFrame({"full_name": [s + " - General" for s in all_positions],
                               "short_name": [s + "-Ge" for s in all_positions],
                               "positions": all_positions})
    ret_positions = pd.concat([positions_internal, fake_roles], ignore_index=True)

    # Handle weights dataframe
    # Dict used to get the average of the normal/green/blue attributes
    weights_general_dict = {"green": 1, "blue": 0, "normal": -1}
    weights_general_dict_reverse = {v: k for k, v in weights_general_dict.items()}
    weights_numeric = weights_internal.replace(weights_general_dict)
    for pos in fake_roles["positions"]:
        relevant_roles = filter_positions_df_by_position(positions_internal, pos)["full_name"]
        relevant_weights = weights_numeric.loc[:, relevant_roles]
        relevant_weights = relevant_weights.mean(axis=1).round(0)
        relevant_weights.replace(weights_general_dict_reverse, inplace=True)
        relevant_weights.name = fake_roles.loc[fake_roles["positions"] == pos, "full_name"].iloc[0]
        weights_internal = pd.concat([weights_internal, relevant_weights], axis=1)
    return ret_positions, weights_internal

# For each player role, gets the weights_internal for that role, multiply by each player
# attributes_internal and divide by sum of the weights, so we get a result between 0 and 20
def calculate_overall(positions_internal, weights_internal, players_internal,
                      player_roles_internal):
    ret = pd.DataFrame()
    for role in player_roles_internal:
        role_full_name = positions_internal.loc[
            positions_internal["short_name"] == role, "full_name"].iloc[0]
        filtered_weights = weights_internal.loc[
            ~pd.isna(weights_internal[role_full_name]), role_full_name]
        total = players_internal[filtered_weights.keys()].multiply(filtered_weights, axis="columns")
        total = total.apply(sum, axis="columns") / filtered_weights.sum()
        ret[role] = total
    return ret

# Function to get the top three best overalls and their respective role names
# TODO: Improve this method by implementing it as a loop
def get_top_overalls(row):
    sorted_row = row.sort_values(ascending=False)
    first_overall_col = sorted_row.index[0]
    second_overall_col = sorted_row.index[1]
    third_overall_col = sorted_row.index[2]
    first_overall = row[first_overall_col]
    second_overall = row[second_overall_col]
    third_overall = row[third_overall_col]
    if first_overall_col.startswith("norm - "):
        column_names = ["1st_norm_over", "1st_norm_over_role", "2nd_norm_over",
                        "2nd_norm_over_role", "3rd_norm_over", "3rd_norm_over_role"]
        first_overall_col = first_overall_col.replace("norm - ", "")
        second_overall_col = second_overall_col.replace("norm - ", "")
        third_overall_col = third_overall_col.replace("norm - ", "")
    else:
        column_names = ["1st_over", "1st_over_role", "2nd_over",
                        "2nd_over_role", "3rd_over", "3rd_over_role"]
    return pd.Series([first_overall, first_overall_col, second_overall, second_overall_col,
                      third_overall, third_overall_col], index=column_names)

# Given all overalls, get the normalized value going from 0 to 100 for each role
def calculate_normalized_overall(players_internal, player_roles_internal):
    max_overalls = players_internal[player_roles_internal].apply(max, axis="rows")
    min_overalls = players_internal[player_roles_internal].apply(min, axis="rows")
    ret = 100 * (players_internal[player_roles_internal] - min_overalls) / (max_overalls -
                                                                             min_overalls)
    # Names of player roles normalized overall columns
    player_roles_norm = "norm - " + player_roles_internal
    ret.columns = player_roles_norm
    return ret

# Remove positions_internal that a player does not play on from a dataframe
def remove_unplayed_positions(row, positions_internal):
    # Get all positions_internal the player does not play on
    not_played_positions = positions_internal[positions_internal["positions"].apply(
        lambda x: all(item not in str(x) for
                       item in row["edited_positions"].split(",")) if not pd.isna(x) else False)]
    # Transform all of those positions_internal to NA
    row[not_played_positions["short_name"]] = pd.NA
    return row

# Returns the best player, the role and overall value of a dataframe containing only overalls
def find_max_value_location(df):
    # Replace NaN values with a very low number
    df_filled = df.fillna(-float("inf"))

    # Find the location of the maximum value
    max_location = df_filled.stack().idxmax()
    # Extract idu, role, and maximum value
    idu, role = max_location
    max_value = df.at[idu, role]
    return idu, role, max_value

# Get a clean dataframe removing players_internal from a list of dataframes
def exclude_players(players_internal, players_to_exclude):
    filtered_players = players_internal.copy()
    # If players_internal to exclude is provided, remove these players_internal
    if players_to_exclude is not None:
        for df in players_to_exclude:
            players_to_drop = [index for index in df.index if index in filtered_players.index]
            filtered_players = filtered_players.drop(players_to_drop)
    return filtered_players

# Given a players_internal dataframe returns the best players
# If team formation is provided, 11 players for these positions are returned
def get_top_players(players_internal, positions_internal,
                    players_to_exclude=None, team_formation_internal=None, age=None, num_players=5,
                    reference_team=None, knowledge=None):
    filtered_players = exclude_players(players_internal, players_to_exclude)
    # Filter by age
    if age is not None:
        filtered_players = filtered_players[filtered_players["Idade"] <= age]
    # Filter by knowledge
    if knowledge is not None:
        filtered_players = filtered_players[filtered_players["Nív. Conh."] <= knowledge]

    if isinstance(reference_team, pd.DataFrame):
        # If comparing to a team, list all positions_internal we need to find
        # players_internal better than the list
        missing_positions = list(set(reference_team["s_position"]))
    elif team_formation_internal is not None:
        # If mounting a team, get all positions_internal we need to fill
        missing_positions = team_formation_internal.copy()
    # If position based, remove all overalls from positions_internal the each player do not play on
    if isinstance(reference_team, pd.DataFrame) or team_formation_internal is not None:
        filtered_players = filtered_players.apply(remove_unplayed_positions, axis=1,
                                                  positions_internal=positions_internal)
    # Get overalls
    all_general_roles = [p + "-Ge" for p in get_all_possible_positions(players_internal)]
    filtered_players = filtered_players[all_general_roles]
    ret = []
    add_player_to_ret = True

    while True:
        if team_formation_internal is not None or isinstance(reference_team, pd.DataFrame):
            # Keep only the general roles of team formation missing positions
            missing_general_roles = [p + "-Ge" for p in set(missing_positions)]
            filtered_players = filtered_players[missing_general_roles]

        if (filtered_players.shape[0] == 0 or filtered_players.shape[1] == 0 or
            filtered_players.isna().all().all()):
            break

        # Get best player for all of missing positions_internal
        idu, role, max_value = find_max_value_location(filtered_players)

        # If position based, check in which position player will play
        if team_formation_internal is not None or isinstance(reference_team, pd.DataFrame):
            selected_position = positions_internal.loc[positions_internal["short_name"] == role,
                                                       "positions"].iloc[0]
        else:
            selected_position = players_internal.loc[idu, "edited_positions"]

        if isinstance(reference_team, pd.DataFrame):
            # If we are comparing to a team, now it is time to decide if we should add this best
            # player to the list or not
            # If player is better than the reference team, add to ret
            if max_value > min(reference_team.loc[
                reference_team["s_position"] == selected_position, "s_over"]):
                add_player_to_ret = True
            else:
                # Case our best player is not good enough for that position, remove the position
                # from the list as we are giving up on it
                add_player_to_ret = False
                missing_positions.remove(selected_position)

        if add_player_to_ret:
            # Add player to ret dataframe and remove the player from possible players_internal
            # to be used
            ret.append([idu, selected_position, role, max_value])
            filtered_players = filtered_players.drop(idu)

        if team_formation_internal is not None:
            # If mounting a team, remove position from from missing positions_internal
            missing_positions.remove(selected_position)

        # Should we leave?
        if team_formation_internal is not None or isinstance(reference_team, pd.DataFrame):
            if len(missing_positions) == 0:
                break
        else:
            num_players -= 1
            if num_players == 0:
                break

    ret = pd.DataFrame(ret, columns=["IDU", "s_position", "s_role", "s_over"])
    ret = ret.set_index("IDU")
    return ret

# Given a set of filtered dataframes and tags pairs, returns a column of the original original
# dataframe containing these tags for the relevant row contained in the filtered dataframes
def get_analysis_status(players_internal, df_and_tag):
    players_copy = pd.DataFrame(index=players_internal.index, columns=["analysis_status"])

    for df, tag in df_and_tag:
        players_copy.loc[players_copy.index.isin(df.index), "analysis_status"] = tag

    return players_copy["analysis_status"]

# Function to convert percentage strings to float
def handle_percentage_values(percentage_str):
    # Remove the percentage sign and convert to float
    return float(percentage_str.strip("%")) / 100

# Organize dataframe so it has a nice format for output
def make_df_printable(full_df, all_notes, filtered_df=None, positions_internal=None, pos=None):
    # Relevant columns for all
    col_all_players = ["s_position", "s_over", "Nome", "analysis_status", "Posição", "Idade",
                       "Clube", "Nac", "Valor", "Preço Exigido", "Salário", "club_status", "Expira",
                       "Pé Preferido", "Altura", "Peso", "Personalidade", "Nív. Conh.",
                       "Situação de Transferência", "Empréstimo"]
    # Relevant columns for non per position list
    col_non_pos = ["1st_over", "1st_over_role", "2nd_over", "2nd_over_role",
                   "3rd_over", "3rd_over_role", "1st_norm_over", "1st_norm_over_role",
                   "2nd_norm_over", "2nd_norm_over_role", "3rd_norm_over", "3rd_norm_over_role"]

    # If coaches dataframe
    if all(full_df["df_type"] == "coaches"):
        # Relevant columns for coaches dataframe
        col_coaches = ["Nome", "Função", "Função Preferida", "Idade", "Clube", "Nac", "Salário",
                       "Expira", "Tipo de Treino", "Qualificações de Treinador", "Personalidade",
                       "Jovens"]
        col_coaches = col_coaches + list(
            positions_internal.loc[positions_internal["positions"].isna(), "short_name"].unique())
        columns = col_coaches + col_non_pos
        ret = full_df
    # If per position dataframe, filter players dataframe
    elif pos is not None:
        ret = filter_players_by_position(full_df, pos)
        # Order by position general attribute
        ret = ret.sort_values(pos + "-Ge", ascending=False)
        col_pos = list(filter_positions_df_by_position(positions_internal, pos)["short_name"])
        columns = list(pos + "-Ge") + col_all_players + col_pos + ["norm - " + c for c in col_pos]
    # If top_players dataframe, get info from full players dataframe
    elif "s_position" in filtered_df.columns:
        ret = pd.merge(filtered_df, full_df, how="inner", left_index=True, right_index=True)
        # Order by team formation
        ret = sort_dataframe_by_custom_order(ret, "s_position", TEAM_FORMATION)
        columns = col_all_players + col_non_pos
    else:
        ret = filtered_df
        columns = col_all_players + col_non_pos
    # Get columns available on filtered_df
    ret = ret[[c for c in columns if c in ret.columns]]
    # Add notes column
    notes_column = all_notes.reindex(ret.index)
    notes_column = notes_column.fillna('')
    ret = notes_column.join(ret)
    return ret

# Used to sort players dataframe on team formation order
def sort_dataframe_by_custom_order(df, column_name, custom_order):
    # Append items from df that are not on custom_order to the end of custom_order
    extra_items = list(df[column_name].unique())
    extra_items.sort()
    custom_order.extend(extra_items)
    # Remove duplicates of custom_order
    custom_order = list(dict.fromkeys(custom_order))
    # Create a categorical data type with the custom order
    custom_order_cat = pd.Categorical(df[column_name], categories=custom_order, ordered=True)
    # Sort the dataframe based on the custom order
    sorted_df = df.assign(**{column_name: custom_order_cat}).sort_values(column_name)
    return sorted_df

def team_summary_helper(names, results, base_name, values):
    names.append(base_name + "_mean")
    results.append(values.mean())
    names.append(base_name + "_max")
    results.append(values.max())
    names.append(base_name + "_min")
    results.append(values.min())
    return names, results

def create_team_summary(outputs, players_internal):
    row_names = []
    row_values = []
    for out in outputs:
        if out[1] in ['18_1st', '21_1st', '1st', '2nd']:
            # row_names contains metrics names and row_values its contents
            row_names.append(out[1] + "_number_of_players")
            row_values.append(len(out[0]))
            [row_names, row_values] = team_summary_helper(
                row_names, row_values, out[1] + "_s_over", out[0]['s_over'])
            [row_names, row_values] = team_summary_helper(
                row_names, row_values, out[1] + "_1st_norm_over",
                players_internal.loc[out[0].index]['1st_norm_over'])
            [row_names, row_values] = team_summary_helper(
                row_names, row_values, out[1] + "_age",
                players_internal.loc[out[0].index]['Idade'])
    team_summary = {IN_GAME_DATE: row_values}
    team_summary = pd.DataFrame(team_summary, index=row_names, )
    team_summary = team_summary.rename_axis('category')
    return team_summary

def get_next_available_file_name(base_name, extension):
    all_files = os.listdir(OUTPUT_DIR)
    i = 1
    while True:
        ret = base_name + str(i) + extension
        if ret not in all_files:
            break
        i = i + 1
    return ret

# Save xlsx results for processed data and a csv for raw data
def save_results(outputs, players_internal, coaches_internal, positions_internal):
    # Before writting, try to read the last results file and get previous team summary
    # and all player notes
    all_files = os.listdir(OUTPUT_DIR)
    results_files = [file for file in all_files if
                     (file.endswith('.xlsx') and file.startswith('results'))]
    most_recent_file = max(results_files,
                           key=lambda x: os.path.getmtime(os.path.join(OUTPUT_DIR, x)))

    # Get the tab names we need to read (all but raw)
    excel = load_workbook(os.path.join(OUTPUT_DIR, most_recent_file), read_only=True,
                          keep_links=False)
    tab_names = excel.sheetnames
    tab_names = [t for t in tab_names if t != "raw"]

    # Read each tab as a separate dataframe
    all_tabs = pd.read_excel(os.path.join(OUTPUT_DIR, most_recent_file), sheet_name=tab_names)
    player_tabs = []
    for sheet_name, df in all_tabs.items():
        if sheet_name == "summary":
            team_summary = df
        else:
            player_tabs.append(df)

    # Read all notes and create a per player notes df
    all_player_tabs_df = pd.concat(player_tabs)
    all_player_tabs_df['note'] = all_player_tabs_df['note'].fillna("")
    all_player_tabs_df['note'] = all_player_tabs_df['note'].astype(str)
    all_notes = all_player_tabs_df.groupby('IDU')['note'].apply(
        lambda x: ''.join(set(filter(None, x)))).reset_index()
    all_notes = all_notes.set_index("IDU")

    # Create team summary
    team_summary = team_summary.set_index("category")
    new_team_summary = create_team_summary(outputs, players_internal)
    new_team_summary = team_summary.join(new_team_summary, rsuffix="_1")

    # Define output file names
    results_name = get_next_available_file_name("results_", ".xlsx")
    raw_players_name = get_next_available_file_name("raw_players_", ".csv")
    raw_coaches_name = get_next_available_file_name("raw_coaches_", ".csv")

    # Save results xlsx
    with pd.ExcelWriter(os.path.join(OUTPUT_DIR, results_name)) as writer: # pylint: disable=abstract-class-instantiated
        # First add team summary
        new_team_summary.to_excel(writer, sheet_name="summary", index=True, float_format="%.2f")
        # Add all players tabs
        for out in outputs:
            make_df_printable(players_internal, all_notes, filtered_df=out[0]).to_excel(
                writer, sheet_name=out[1], index=True, float_format="%.2f")
        # Add all positions tabs
        for pos in get_all_possible_positions(players_internal):
            make_df_printable(players_internal, all_notes, positions_internal=positions_internal,
                              pos=pos).to_excel(writer, sheet_name=pos, index=True,
                                                float_format="%.2f")
        # Add coaches tab
        make_df_printable(
            coaches_internal, all_notes, positions_internal=positions_internal).to_excel(
                writer, sheet_name="coach", index=True, float_format="%.2f")
    # Also save raw data as a csv
    players_internal.to_csv(os.path.join(OUTPUT_DIR, raw_players_name),
                            index=True, float_format="%.2f", encoding='utf-8-sig')
    coaches_internal.to_csv(os.path.join(OUTPUT_DIR, raw_coaches_name),
                            index=True, float_format="%.2f", encoding='utf-8-sig')

################## Main ##################
def main():
    # Get current directory and read attributes weights CSVs
    curr_dir = os.path.dirname(os.path.abspath(__file__))
    attributes = pd.read_csv(os.path.join(curr_dir, "relevant_attributes", "attributes.csv"))
    positions = pd.read_csv(os.path.join(curr_dir, "relevant_attributes", "positions.csv"))
    weights = pd.read_csv(os.path.join(curr_dir, "relevant_attributes", "weights.csv"),
                          index_col="short_name")

    # Read data directory, get all files, and order from old to new
    players, coaches = read_data()

    # Drop duplicates considering unique ID and keeping most recent date
    players = players.sort_values("timestamp").drop_duplicates(["IDU"], keep="last")
    coaches = coaches.sort_values("timestamp").drop_duplicates(["IDU"], keep="last")
    players = players.set_index("IDU")
    coaches = coaches.set_index("IDU")

    # Handle player names removing country name
    players["Nome"] = players["Nome"].str.split(" - ").str[0]

    # Handle percentage values
    players["Nív. Conh."] = players["Nív. Conh."].apply(handle_percentage_values)

    # Handle value column
    players["Valor"] = players["Valor"].fillna(players["Valor.1"])
    players.drop(columns=["Valor.1"], inplace=True)

    # Handle loan and transfer status text
    players["Situação de Transferência"] = players["Situação de Transferência"].replace(
        {"Colocado na lista de transferências": "Listado"})
    players["Empréstimo"] = players["Empréstimo"].replace(
        {"Colocado na lista de transferências": "Listado"})

    # Get person club status
    players["club_status"] = players.apply(get_person_club_status, axis=1)
    coaches["club_status"] = coaches.apply(get_person_club_status, axis=1)

    # Parse players positions - Go from "MD, M/MO (DC), PL (C)" to MD,M(D),M(C),MO(D),MO(C),PL(C)
    players["edited_positions"] = players["Posição"].apply(parse_position)

    # Classify player/coach on attribute availability - complete, complete but range,
    # incomplete, all incomplete
    players["data_completion"] = players.apply(lambda x: check_data_completion(x, attributes),
                                               axis=1)
    coaches["data_completion"] = coaches.apply(lambda x: check_data_completion(x, attributes),
                                               axis=1)

    # For coaches, there is no incomplete attributes. Keep only complete rows
    coaches = coaches[coaches["data_completion"] == "complete"]

    # Get average attribute value per position
    average_attributes_pos = get_average_attribute_per_position(players, attributes)

    # Handle missing data for all players that are "complete but range" or "incomplete"
    players = players.apply(lambda x: handle_missing_attributes(x, attributes,
                                                                average_attributes_pos), axis=1)

    # Convert attributes to numeric
    players = convert_df_columns_to_numeric(
        players, attributes.loc[attributes["player_coach"] == "player", "short_name"])
    coaches = convert_df_columns_to_numeric(
        coaches, attributes.loc[attributes["player_coach"] == "coach", "short_name"])

    # Add general roles to positions and weights df
    positions, weights = create_general_roles(players, positions, weights)

    # Replace the attribute colors with weight values
    player_roles_full_name = positions.loc[positions["positions"].notna(), "full_name"].unique()
    coach_roles_full_name = positions.loc[positions["positions"].isna(), "full_name"].unique()
    weights[player_roles_full_name] = weights[player_roles_full_name].replace(WEIGHTS_DICT_PLAYERS)
    weights[coach_roles_full_name] = weights[coach_roles_full_name].replace(WEIGHTS_DICT_COACHES)

    # Calculate overall per role
    player_roles = positions.loc[positions["positions"].notna(), "short_name"].unique()
    players = pd.concat([players, calculate_overall(positions, weights, players, player_roles)],
                        axis="columns")
    coach_roles = positions.loc[positions["positions"].isna(), "short_name"].unique()
    coaches = pd.concat([coaches, calculate_overall(positions, weights, coaches, coach_roles)],
                        axis="columns")

    # Get the top three overalls
    result = players[player_roles].apply(get_top_overalls, axis="columns")
    players = pd.concat([players, result], axis="columns")
    result = coaches[coach_roles].apply(get_top_overalls, axis="columns")
    coaches = pd.concat([coaches, result], axis="columns")

    # Calculate normalized overalls
    players = pd.concat([players, calculate_normalized_overall(players, player_roles)],
                        axis="columns")
    coaches = pd.concat([coaches, calculate_normalized_overall(coaches, coach_roles)],
                        axis="columns")

    # Get the top three normalized overalls
    result = players["norm - " + player_roles].apply(get_top_overalls, axis="columns")
    players = pd.concat([players, result], axis="columns")
    result = coaches["norm - " + coach_roles].apply(get_top_overalls, axis="columns")
    coaches = pd.concat([coaches, result], axis="columns")

    # Get a de-fragmented dataframe for players for performance
    players = players.copy()
    coaches = coaches.copy()

    # Get best 11 and best 2nd team
    best_eleven = get_top_players(players[players["club_status"] == "my_club"], positions,
                                  team_formation_internal=TEAM_FORMATION)
    best_2nd_team = get_top_players(players[players["club_status"] == "my_club"], positions,
                                    [best_eleven], TEAM_FORMATION)
    # Get lent players good enough for best eleven and best 2nd team
    lent_best_eleven = get_top_players(players[players["club_status"] == "my_club_lent"],
                                       positions, reference_team=best_eleven)
    lent_best_2nd_team = get_top_players(players[players["club_status"] == "my_club_lent"],
                                         positions, [lent_best_eleven],
                                         reference_team=best_2nd_team)
    # Get best 11 under 21 and 18
    under_21_best_eleven = get_top_players(
        players[players["club_status"].isin(["my_club_lent", "my_club"])], positions,
        [best_eleven, best_2nd_team, lent_best_eleven, lent_best_2nd_team],
        team_formation_internal=TEAM_FORMATION, age=21)
    under_18_best_eleven = get_top_players(
        players[players["club_status"].isin(["my_club_lent", "my_club"])], positions,
        [best_eleven, best_2nd_team, lent_best_eleven, lent_best_2nd_team, under_21_best_eleven],
        team_formation_internal=TEAM_FORMATION, age=18)
    # Get the bad players
    bad_players_my_team = exclude_players(players[players["club_status"] == "my_club"],
                                  [best_eleven, best_2nd_team, under_21_best_eleven,
                                   under_18_best_eleven])
    # Get best lent players
    bad_lent = exclude_players(players[players["club_status"] == "my_club_lent"],
                               [lent_best_eleven, lent_best_2nd_team, under_21_best_eleven,
                                under_18_best_eleven])

    # Handle players with low knowledge
    # TODO: Some classifications are not used anymore. Remove
    no_knowledge = players[players["data_completion"] == "all_incomplete"]
    low_knowledge_best_2nd_team = get_top_players(
        players[players["df_type_detailed"] == "search_players"], positions,
        [no_knowledge], reference_team=best_2nd_team, knowledge=0.3)
    low_knowledge_best_2nd_team_under_21 = get_top_players(
        players[players["df_type_detailed"] == "search_players"], positions,
        [no_knowledge, low_knowledge_best_2nd_team], reference_team=under_21_best_eleven,
        age=21, knowledge=0.3)
    low_knowledge_best_top_5 = get_top_players(
        players[players["df_type_detailed"] == "search_players"], positions,
        [no_knowledge, low_knowledge_best_2nd_team, low_knowledge_best_2nd_team_under_21],
        knowledge=0.3)
    low_knowledge_best_top_5_under_21 = get_top_players(
        players[players["df_type_detailed"] == "search_players"], positions,
        [no_knowledge, low_knowledge_best_2nd_team, low_knowledge_best_2nd_team_under_21,
         low_knowledge_best_top_5], age=21, knowledge=0.3)
    low_knowledge_bad_players = exclude_players(
        players[(players["df_type_detailed"] == "search_players") & (players["Nív. Conh."] <= 0.3)],
        [no_knowledge, low_knowledge_best_2nd_team, low_knowledge_best_2nd_team_under_21,
         low_knowledge_best_top_5, low_knowledge_best_top_5_under_21])
    low_knowledge_suggest_scout = exclude_players(
        players[(players["df_type_detailed"] == "search_players") & (players["Nív. Conh."] <= 0.3)],
        [low_knowledge_bad_players])
    # Players with knowledge
    # TODO: Some classifications are not used anymore. Remove
    hire_best_eleven = get_top_players(players[players["df_type_detailed"] == "search_players"],
                                       positions,
                                       [no_knowledge, low_knowledge_bad_players,
                                        low_knowledge_suggest_scout], reference_team=best_eleven)
    hire_best_2nd_team = get_top_players(players[players["df_type_detailed"] == "search_players"],
                                         positions,
                                         [no_knowledge, low_knowledge_bad_players,
                                          low_knowledge_suggest_scout, hire_best_eleven],
                                          reference_team=best_2nd_team)
    hire_best_under_21 = get_top_players(players[players["df_type_detailed"] == "search_players"],
                                         positions,
                                         [no_knowledge, low_knowledge_bad_players,
                                          low_knowledge_suggest_scout, hire_best_eleven,
                                          hire_best_2nd_team],
                                          reference_team=under_21_best_eleven, age=21)
    non_team_formation_positions = list(set(get_all_possible_positions(players)) -
                                        set(TEAM_FORMATION))
    hire_best_other_positions = get_top_players(
        players[players["df_type_detailed"] == "search_players"],
        positions,
        [no_knowledge, low_knowledge_bad_players, low_knowledge_suggest_scout, hire_best_eleven,
         hire_best_2nd_team, hire_best_under_21],
         team_formation_internal=non_team_formation_positions)
    bad_hire_players = exclude_players(players[players["df_type_detailed"] == "search_players"],
                                       [no_knowledge, low_knowledge_bad_players,
                                        low_knowledge_suggest_scout, hire_best_eleven,
                                        hire_best_2nd_team, hire_best_under_21,
                                        hire_best_other_positions])

    # Add a tags column to the players dataframe
    players["analysis_status"] = get_analysis_status(
        players, [(best_eleven, "best_11"), (best_2nd_team, "best_2nd_team"),
                  (lent_best_eleven, "lentbest_11"), (lent_best_2nd_team, "lent_best_2nd_team"),
                  (under_21_best_eleven, "under_21_best_11"),
                  (under_18_best_eleven, "under_18_best_11"),
                  (bad_players_my_team, "bad_players_my_team"),
                  (bad_lent, "bad_lent"), (no_knowledge, "no_knowledge"),
                  (low_knowledge_bad_players, "low_knowledge_bad_players"),
                  (low_knowledge_suggest_scout, "low_knowledge_suggest_scout"),
                  (hire_best_eleven, "hire_best_11"), (hire_best_2nd_team, "hire_best_2nd_team"),
                  (hire_best_under_21, "hire_best_under_21"),
                  (hire_best_other_positions, "hire_best_other_positions"),
                  (bad_hire_players, "bad_hire_players")])

    save_results([(best_eleven, "1st"),
                  (best_2nd_team, "2nd"),
                  (under_21_best_eleven, "21_1st"),
                  (under_18_best_eleven, "18_1st"),
                  (lent_best_eleven, "lent_1st"),
                  (lent_best_2nd_team, "lent_2nd"),
                  (bad_players_my_team, "bad"),
                  (bad_lent, "bad_lent")], players, coaches, positions)

    print("end")

if __name__ == "__main__":
    main()

#weights = convertDfColumnsToNumeric(weights, [c for c in weights.columns if c != "short_name"])

#i = 0
#for a in players["Pas"]:
#    if type(a) is np.ndarray:
#        print(i)
#    i = i + 1

# check unique content per column
#for key in players.keys():
#    print(key)
#    print(len(players[key].unique()))
#    print(players[key].unique())
#    print("\n\n\n\n\n\n\n______________________________________")
#for key in coaches.keys():
#    print(key)
#    print(len(coaches[key].unique()))
#    print(coaches[key].unique())
#    print("\n\n\n\n\n\n\n______________________________________")
